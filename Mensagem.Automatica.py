from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from openpyxl import Workbook, load_workbook
from pathlib import Path 
from urllib.parse import quote
from time import sleep
from datetime import datetime, timedelta
import os
import sys
import locale
import platform
import subprocess
import requests
import json


load_da_pagina = '//*[@id="app"]/div/div[2]/div[3]/header/header/div/span/div/span/div[2]/div/span'
botao_de_envio = '//*[@id="main"]/footer/div[1]/div/span[2]/div/div[2]/div[2]/button/span'
botao_invalido = '//*[@id="app"]/div/span[2]/div/span/div/div/div/div/div/div[2]/div/button' 



def versao():
    versao = "v1.10"


    caminho_arquivo_versao = "versao.txt"

    with open(caminho_arquivo_versao, "w") as arquivo_versao:
        arquivo_versao.write(versao)

    with open('versao.txt', 'r', encoding='utf-8') as arquivo:
        texto = arquivo.read()
        
    return texto

def editar_mensagem():
    msg = Path('Mensagem.txt')

    if not msg.exists():

        with open('Mensagem.txt', 'w') as arquivo:
            pass

def conferir_versão(versao_atual):
    try:
        response = requests.get('https://api.github.com/repos/SethiCoel/atualizacao-bot-wpp/releases/latest')
        latest_release = json.loads(response.text)
        ultima_versao = latest_release['tag_name']


        if ultima_versao != versao_atual:
            print('Nova versão disponível!')

            caminho_executavel = 'update/update.exe'
            command = f'start {caminho_executavel}'
            subprocess.Popen(command, shell=True)
            sys.exit()
        

    except Exception as error:
        print(error)

def converter_xls_em_xlsx():
    try:    
        import jpype     
        import asposecells     
        jpype.startJVM() 
        from asposecells.api import Workbook

        diretorio_atual = os.getcwd()

        arquivos = os.listdir(diretorio_atual)

        planilha_xls = None

        for arquivo in arquivos:
            if arquivo.endswith(".xls"):
                planilha_xls = os.path.join(diretorio_atual, arquivo)
                break

        if planilha_xls:
            workbook = Workbook(f"{planilha_xls}")
            workbook.save("Planilha MK-AUTH/Planilha MK-AUTH.xlsx")
            jpype.shutdownJVM()
            os.remove(planilha_xls)

        else:
            pass
    except Exception as error:
        os.system('cls')
        print('Aconteceu um erro ->',error)

def apagar_cache():
    #verifica se o sistema operacional é windows
    if platform.system() == 'Windows':
        planilha = Path('Não Enviados/Planilha de Reenvio.xlsx')

        #se a data de criação do arquivo for diferete de hj ele será apagado
        if planilha.exists():
            info_arquivo = os.stat(planilha)
            data_criacao = datetime.fromtimestamp(info_arquivo.st_ctime).date()
            data_hoje = datetime.now().date()

            if data_criacao != data_hoje:
                os.remove(planilha)
        else:
            pass

    else:
        pass

def planinha_atualizada():

    os.system('cls')

    pasta = Path('Planilha MK-AUTH')
    
    if not pasta.exists():
        pasta.mkdir()
        print('Pasta não encontrada')
        sleep(1)

        print('Criando nova pasta...')
        sleep(1)

        print('Pasta criada com sucesso!')
        sleep(1)
        pass


    diretorio = "Planilha MK-AUTH/"
    arquivo_att = Path('Planilha Atualizada.xlsx')

    # Listar todos os arquivos no diretório
    arquivos = os.listdir(diretorio)

    arquivo_xlsx = None
    
    for arquivo in arquivos:
        if arquivo.endswith(".xlsx"):
            arquivo_xlsx = os.path.join(diretorio, arquivo)
            break


    if arquivo_xlsx:

        try:

            if not arquivo_att.exists():
                print("Planilha encontrada!")
                sleep(1)
                
                print('Criando uma nova planilha atualizada...')
                sleep(1)

                print('Planilha atualizada criada com sucesso!')
                sleep(1)


            workbook = openpyxl.load_workbook(f'{arquivo_xlsx}')
    
            pagina_clientes = workbook['Sheet1']

            wb = Workbook()
            ws = wb.active

            ws.append(['Nome', 'Número', 'Vencimento'])


            for id, linha in enumerate(pagina_clientes.iter_rows(min_row=3)):
                

                if arquivo_att.exists():
                    pass

                else:

                    print(f'{id}: {linha[1].value}| Número: {linha[15].value} | Vencimento: {linha[25].value}')
                

                nome = linha[1].value
                numero = linha[15].value
                vencimento = linha[25].value
                

                ws.append([f'{nome}', f'{numero}', f'{vencimento}'])
            
            os.system('cls')
            wb.save('Planilha Atualizada.xlsx')

            # Carregando a planilha original e a planilha cópia
            wb_original = load_workbook(f'{arquivo_xlsx}')
            
            wb_copia = load_workbook('Planilha Atualizada.xlsx')

            # Selecionando as folhas ativas
            ws_original = wb_original.active
            ws_copia = wb_copia.active

            ws_copia.column_dimensions['A'].width = 40
            ws_copia.column_dimensions['B'].width = 20
            ws_copia.column_dimensions['C'].width = 15

            wb_copia.save('Planilha Atualizada.xlsx')
            

        except KeyError:
            os.system('cls')
            print('Erro:')
            print('O WorkSheet dentro da planilha do MK-AUTH precisa ser renomeada para Sheet1 ')
            input('Presioner ENTER para fechar.')
            sys.exit()
        
        except ImportError as error:
            print('Erro de importação:',error)

    
    if not arquivo_att.exists():
        os.system('cls')
        cor('Nenhuma Planilha encontrada!', 'vermelho')
        input('Pressione ENTER para fechar')
        sys.exit()


    else:
        if pasta.exists():
            menu()

def planilha_de_reenvio():
    pasta = Path('Não Enviados')
    if not pasta.exists():
        pasta.mkdir()

    planilha = Path('Não Enviados/Planilha de Reenvio.xlsx')

    if not planilha.exists():
        wb = Workbook()
        ws = wb.active

        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20

        wb.save('Não Enviados/Planilha de Reenvio.xlsx') 

def menu():
    while True:
        os.system('cls')
        print(f'''WhatsApp Bot de Mensagem Automática                                                             versão:{versao()} 

(1) Ativar Mensagem Automática
(2) Tentar Reenviar Mensagens 
(3) Programar o BOT
(4) Verificar mensagem
    
              
Instruções digite -> ajuda              
              
                    ''')

        opcao = input(f'Digite 1 Para Ativar o Bot: ')

        if opcao == 'ajuda':
            os.system('cls')
            ajuda()

        if opcao == '1':
            os.system('cls')
            print('PARA CANCELAR FECHE O APP...')
            sleep(3)            
            mensagem_automatica()
            break
        
        if opcao == '2':
            os.system('cls')
            print('PARA CANCELAR FECHE O APP...')
            sleep(3)  
            reenviar_mensagem()

        if opcao == '3':
            programar()
        
        if opcao == '4':
            os.system('cls')
            mensagem = '''*Mensagem Automática:*

Olá {XX}, seu boleto vence dia {XX} (amanhã).'''
                
            with open('mensagem.txt', 'r', encoding='utf-8') as arquivo:
                texto = arquivo.read()
                nao_modificavel = f'\033[0;36;40m{mensagem}\033[m'
                modificavel = f'\033[0;32;40m{texto}\033[m'
                cor('(Apenas o texto em verde pode ser modificado)\n','vermelho')
                print(nao_modificavel,modificavel)
            
            input('\n\nPressione ENTER para voltar')
        
        else:
            continue

def mensagem_automatica():
    send_list = []
    notsend_list = []

    current_directory = os.path.dirname(os.path.abspath(__file__))

    session_data_directory = os.path.join(current_directory, "session_data")

    # Verifica se a pasta de sessão existe, senão a cria
    if not os.path.exists(session_data_directory):
        os.makedirs(session_data_directory)

    chrome_options = Options()

    chrome_options.add_argument(f'--user-data-dir={session_data_directory}')
    chrome_options.add_argument('--profile-directory=Default')
    chrome_options.add_argument('--disable-extensions')
    chrome_options.add_argument('--disable-popup-blocking')

    chrome_options.add_experimental_option("detach", True)
    servico = Service(ChromeDriverManager().install())
    
    navegador = webdriver.Chrome(service=servico, options=chrome_options)
    navegador.get(f'https://web.whatsapp.com')
    
    
    WebDriverWait(navegador, 6000).until(
        EC.element_to_be_clickable((By.XPATH, load_da_pagina))
    )

    sleep(3)
    navegador.quit()

    workbook = openpyxl.load_workbook('Planilha Atualizada.xlsx')
    pagina_clientes = workbook['Sheet']

    # Extrai todos os dados da planilha cópia para o envio das mensagens
    for linha in pagina_clientes.iter_rows(min_row=2):
        

        nome = linha[0].value
        telefone = linha[1].value
        vencimento = linha[2].value

        if vencimento is not None:
            
            data_antecipada = timedelta(days=int(vencimento)) - timedelta(days=1)        
            data_atual = datetime.now().day

            # Faz a verificação da data, caso seja um dia antes do vencimento ele enviará a mensagem, senão irá ignorar
            if data_antecipada.days == data_atual:

                # Caso o número seja vazio ele é alertado e registrado na planilha "Planilha de Reenvio"
                if telefone is None or telefone == '' or telefone == 'None':
                    planilha_de_reenvio()

                    msg = f"Mensagem não enviada para {nome}"

                    notsend_list.append(f'{msg:<80} {"- Sem Número":>30}')

                    workbook = load_workbook('Não Enviados/Planilha de Reenvio.xlsx')
                    sheet = workbook.active
                    
                    linha = 1
                    coluna = 1
                    
                    while sheet.cell(row=linha, column=coluna).value is not None:
                        linha += 1

                    
                    dados = [f'{nome}', f'', f'{vencimento}', 'Sem Número']
                    
                    for col, dado in enumerate(dados, start=1):
                
                        sheet.cell(row=linha, column=col, value=dado)
                    
                    workbook.save('Não Enviados/Planilha de Reenvio.xlsx')
                    continue
                

                # Mensagem de exemplo que os clientes receberá contendo o nome e o vencimento. Sendo possível a troca a mensagem para a que mais agradar
                mensagem = f'''*Mensagem Automática:*

Olá {nome.title()}, seu boleto vence dia {vencimento} (amanhã).'''
                
                with open('mensagem.txt', 'r', encoding='utf-8') as arquivo:
                    texto = arquivo.read()
                    msg = f'{mensagem} {texto}'

                try:
                    navegador = webdriver.Chrome(service=servico, options=chrome_options)
                    navegador.get(f'https://web.whatsapp.com/send?phone={telefone}&text={quote(msg)}')


                    WebDriverWait(navegador, 6000).until(
                        EC.element_to_be_clickable((By.XPATH, load_da_pagina))
                    )
                                 
                    sleep(3)

                    numero_invalido = WebDriverWait(navegador, 3).until(
                        EC.element_to_be_clickable((By.XPATH, botao_invalido))
                    )
                    
                    planilha_de_reenvio()
                    msg1 = f"Mensagem não enviada para {nome}"

                    notsend_list.append(f'{msg1:<80} {"- Número Inválido":>30}')
                    
                    workbook = load_workbook('Não Enviados/Planilha de Reenvio.xlsx')
                    sheet = workbook.active
                    
                    linha = 1
                    coluna = 1
                    
                    while sheet.cell(row=linha, column=coluna).value is not None:
                        linha += 1

                    dados = [f'{nome}', f'{telefone}', f'{vencimento}', 'Número Inválido']
                    
                    for col, dado in enumerate(dados, start=1):
                
                        sheet.cell(row=linha, column=col, value=dado)
                    
                    workbook.save('Não Enviados/Planilha de Reenvio.xlsx')
                    


                except Exception as error:
                   
                    botao_enviar = WebDriverWait(navegador, 6000).until(
                        EC.element_to_be_clickable((By.XPATH, botao_de_envio))
                    )
                    
                    botao_enviar.click()

                    send_list.append(f'Mensagem enviada com sucesso para {nome}')
                    

                finally:
                    sleep(5)
                    navegador.quit()
    
    locale.setlocale(locale.LC_ALL, 'pt_BR.utf-8')
    data_atual = datetime.now()
    dia_por_extenso = data_atual.strftime('%A')
    
    if dia_por_extenso == 'domingo':
        sleep(2)
        os.system("shutdown /s /t 1")
    
    os.system('cls')
    print('Resultado das mensagens:')
    for pessoas in send_list:
        cor(pessoas,'verde')

    for pessoas in notsend_list:
        cor(pessoas, 'vermelho')

    
    cor('\nExecução Finalizada!', 'azul')
    input('\nPressione ENTER para voltar')
    menu()

def reenviar_mensagem():
    pasta = Path('Não Enviados')
    
    if not pasta.exists():
        os.system('cls')
        print('Não tem mensagens a ser reenviadas no momento.')
        input('Presione ENTER para voltar')
        main()

    send_list = []
    notsend_list = []

    current_directory = os.path.dirname(os.path.abspath(__file__))

    session_data_directory = os.path.join(current_directory, "session_data")

    chrome_options = Options()

    chrome_options.add_argument(f'--user-data-dir={session_data_directory}')
    chrome_options.add_argument('--profile-directory=Default')
    chrome_options.add_argument('--disable-extensions')
    chrome_options.add_argument('--disable-popup-blocking')

    chrome_options.add_experimental_option("detach", True)
    servico = Service(ChromeDriverManager().install())

    navegador = webdriver.Chrome(service=servico, options=chrome_options)
    navegador.get(f'https://web.whatsapp.com')

    WebDriverWait(navegador, 6000).until(
        EC.element_to_be_clickable((By.XPATH, load_da_pagina))
    )

    sleep(3)
    navegador.quit()

    workbook = openpyxl.load_workbook('Não Enviados/Planilha de Reenvio.xlsx')
    pagina_clientes = workbook['Sheet']

    delete_rows = []
    
    # Extrai todos os dados da planilha cópia para o envio das mensagens
    for index, linha in enumerate(pagina_clientes.iter_rows(min_row=1)):
        

        nome = linha[0].value
        telefone = linha[1].value
        vencimento = linha[2].value
        
        if vencimento is not None:
            
            data_antecipada = timedelta(days=int(vencimento)) - timedelta(days=1)        
            data_atual = datetime.now().day

            # Faz a verificação da data, caso seja um dia antes do vencimento ele enviará a mensagem, senão irá ignorar
            if data_antecipada.days == data_atual:

                if telefone != 'Número Inválido' and telefone != 'Sem Número' and telefone is not None:
                    delete_rows.append(index + 1)
                
                # Caso o número seja vazio ele é alertado e registrado na planilha "Planilha de Reenvio"
                if telefone == 'Número Inválido' or telefone == 'Sem Número' or telefone is None:
                    
                    msg = f"Mensagem não enviada para {nome}"
                    
                    notsend_list.append(f'{msg:<80} {"- Sem Número":>30}')

                    continue

                # Mensagem de exemplo que os clientes receberá contendo o nome e o vencimento. Sendo possível a troca a mensagem para a que mais agradar
                mensagem = f'''*Mensagem Automática:*

Olá {nome.title()}, seu boleto vence dia {vencimento} (amanhã).'''
                
                with open('mensagem.txt', 'r', encoding='utf-8') as arquivo:
                    texto = arquivo.read()
                    msg = f'{mensagem} {texto}'

                try:
                    navegador = webdriver.Chrome(service=servico, options=chrome_options)
                    navegador.get(f'https://web.whatsapp.com/send?phone={telefone}&text={quote(msg)}')

                    WebDriverWait(navegador, 6000).until(
                        EC.element_to_be_clickable((By.XPATH, load_da_pagina))
                    )

                    sleep(3)

                    numero_invalido = WebDriverWait(navegador, 3).until(
                        EC.element_to_be_clickable((By.XPATH, botao_invalido))
                    )
                    
                    msg1 = f"Mensagem não enviada para {nome}"
                    notsend_list.append(f'{msg1:<80} {"- Número Inválido":>30}')

                    

                except Exception as error:

                    botao_enviar = WebDriverWait(navegador, 6000).until(
                        EC.element_to_be_clickable((By.XPATH, botao_de_envio))
                    )

                    botao_enviar.click()

                    send_list.append(f'Mensagem enviada com sucesso para {nome}')

                    workbook = load_workbook('Não Enviados/Planilha de Reenvio.xlsx')

                    sheet = workbook.active
                    
                    for linha_para_excluir in reversed(delete_rows):
                        sheet.delete_rows(linha_para_excluir)

                    workbook.save('Não Enviados/Planilha de Reenvio.xlsx')
                    
                    

                finally:
                    sleep(3)
                    navegador.quit()

    
    os.system('cls')
    print('Resultado das mensagens:')
    for pessoas in send_list:
        cor(pessoas, 'verde')
    
    for pessoas in notsend_list:
        cor(pessoas, 'vermelho')
    

    
    cor('\nExecução Finalizada!', 'azul')
    input('\nPressione ENTER para voltar')
    
    send_list = []
    notsend_list = []
    
    main()

def programar():
    os.system('cls')

    locale.setlocale(locale.LC_ALL, 'pt_BR.utf-8')
    data_atual = datetime.now()
    amanha = data_atual + timedelta(days=1)
    amanha_extenso = amanha.strftime('%A')

    print('=' * 74)
    print(f'O BOT será programado para funcionar automaticamente amanhã ({amanha_extenso})!')
    print('=' * 74)

    sn = input('\nDeseja confirmar esta ação? S/N: ').upper()

    if sn == 'S' or sn == 'SIM':

        print('Programando Bot...')
        sleep(1)

        print('Bot programado com sucesso!')
        sleep(1)

        print('Iniciando Bot...')
        sleep(1)
        
        dia = []
        
        os.system('cls')
        while True:


            locale.setlocale(locale.LC_ALL, 'pt_BR.utf-8')
            data_atual = datetime.now()
            amanha = data_atual + timedelta(days=1)
            amanha_extenso = amanha.strftime('%A')
            horario_atual = datetime.now().strftime('%H:%M:%S')
                    
            horario_especifico = datetime.strptime("07:00:00", "%H:%M:%S").strftime('%H:%M:%S')
            
            dia_por_extenso = data_atual.strftime('%A')


            dia.append(amanha_extenso)

            if dia_por_extenso == dia[0] and horario_atual >= horario_especifico:
                os.system('cls')
                print('Iniciando...')
                sleep(3)
                mensagem_automatica()
                break
            
            print(f'Assim que for {amanha_extenso} o bot irá iniciar automaticamente as 07:00Hrs da manhã!')
            print('Não feche o programa!')
            print('\nEsperando.')
            sleep(0.5)
            os.system('cls')

            print(f'Assim que for {amanha_extenso} o bot irá iniciar automaticamente as 07:00Hrs da manhã!')
            print('Não feche o programa!')
            print('\nEsperando..')
            sleep(0.5)
            os.system('cls')

            print(f'Assim que for {amanha_extenso} o bot irá iniciar automaticamente as 07:00Hrs da manhã!')
            print('Não feche o programa!')
            print('\nEsperando...')
            sleep(0.5)
            os.system('cls')    

    else:
        menu()

def main():
    versao()
    conferir_versão(versao())
    apagar_cache()
    converter_xls_em_xlsx()
    editar_mensagem()
    planinha_atualizada()
    menu()

def cor(texto, cor=''):
        if cor == 'vermelho':
            cor = '\033[0;31;40m'
       
        if cor == 'verde':
            cor = '\033[0;32;40m'

        if cor == 'amarelo':
            cor = '\033[0;33;40m'
       
        if cor == 'azul':
            cor = '\033[0;34;40m'

        if cor == 'ciano':
            cor = '\033[0;36;40m'
        
        fim = '\033[m'

        print(f'{cor}{texto}{fim}')

def ajuda():

    cor('''\n\n-----------------------------------------------|Passo a Passo|--------------------------------------------------------------------------------------
''','vermelho')
    
    cor('''\n*Para o funcionamento do bot é preciso da PLANILHA contendo todos os dados dos clientes do MK-AUTH.
 
  -> Para baixar a PLANILHA abra o MK-AUTH
  -> Vá para a aba CLIENTES e depois vá em LISTAR TODOS OS CLIENTES
  -> Marque a PRIMEIRA CAIXINHA com o nome LOGIN para selecionar todos os clientes na pagina atual''','verde')
    
    cor('''\n*Após marcar a primeira caxinha aparecerá logo acima outra caixinha com o nome SELECIONAR TUDO DAS X PAGINAS
(X seria o equivalente a quantidade de paginas após a primeira. Exemplo 1.. 2.. 13.. etc)
  
  -> Marque a caixinha com o nome SELECIONAR TUDO para selecionar todos os outros clientes faltando
  -> Desça ate o FINAL da pagina e clique no ULTIMO ICONE a DIREITA (uma prancheta com papel)''','amarelo')
    cor('''\n*Após apertar no icone certo aparecerá uma tela dizendo: SELECIONE AONDE EXPORTAR
terá varios formatos de arquivos mas o bot foi programado para utilizar planilhas.
  
  -> Clique em: PARA PLANILHA EXCEL e espere ele terminar de baixar
  -> Após baixar coloque a planilha na pasta do bot''','azul')
    
    cor('''\n*Após fazer todo esse passo ta finalizado o processo de adicionar os clientes ao bot
agora é so abrir o execeutável "MENSAGEM AUTOMÁTICA" que o programa irá extrair toda a informação da planilha''','ciano')
    
    cor('\n*Pronto agora é so usar o bot.','vermelho')
    
    cor('\n->Para ATUALIZAR a planilha apenas coloque a NOVA PLANILHA dentro da pasta do bot e abra o programa<-','verde')
    
    cor('''\n--------------------------------------------|! Infomações Importantes !|-----------------------------------------------------------''','amarelo')
    
    cor('''\n- É necessário que ative o bot apenas UMA vez por dia para não enviar duas mensagens para os mesmos clientes.
Caso precise enviar uma mensagem para os clientes que não tiveram êxito, seja por estar sem número ou com número inválido
o programa terá um opção especialmente para isso que é a opção (2), dessa forma não acontecerá o problema de mensagem duplicada! ''','azul')
    
    cor('''\n- Os clientes que não tiveram êxito ficarão numa planilha dentro da pasta NÃO ENVIADOS, assim será possível fazer a correção
dos clientes casos seja possível e assim poder utilizar a opção (2) de reenvio.
Os clientes que obtiverem sucesso serão pagados da planilha e ficará apenas os sem êxito, ficando de forma organizada.''','ciano')
    
    cor('''\n- A planilha de REENVIO ficará disponivél por apenas UM dia
após isso ela será apagada automáticamente da pasta Não enviados.''','vermelho')
    
    cor('''\n- É recomendável que atualize a planilha do bot toda vez que houver uma atualização de data, nome e/ou celular dos clientes.
Caso prefir pode atualizar a planilha do bot uma vez ao mês, ja que o cliente receberá a mensagem uma vez no mês.
Para atualizar a planilha é o mesmo do passo a passo acima.''','verde')
    
    cor('''\n- Aos sábados é necessário utilizar a opção (3) após encerrar o expediente, assim o bot irá funcionar no domingo.
ele irá esperar até que seja domingo para começar a funcionar, então tem que ativer e deixar ele ligado.
assim que for domingo 7:00 Horas ele começará o serviço e assim que acabar irá desligar o computador automaticamente.
''','amarelo')

    input('\n\nPressione ENTER para voltar')
    menu()

if __name__=='__main__':
    main()